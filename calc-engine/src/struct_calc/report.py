"""
Report Generator

Outputs structural design reports in Excel/PDF/JSON format.

Excel layout (6 sheets):
    Sheet 1: 設計總覽 - Project overview, design parameters
    Sheet 2: 混凝土用量 - Concrete by fc' and member type
    Sheet 3: 鋼筋用量 - Rebar tonnage and density check
    Sheet 4: 構件清單 - Revit Family/Type list
    Sheet 5: 造價估算 - Cost breakdown
    Sheet 6: 與Revit比對 - (optional) Comparison with actual Revit model
"""

from __future__ import annotations
import json
from pathlib import Path
from dataclasses import asdict
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from .quantity import QuantityTakeoff, StructuralModel
from .cost import CostBreakdown


# ═══════════════════════════════════════════════════════
# JSON EXPORT (always available)
# ═══════════════════════════════════════════════════════

def export_json(
    takeoff: QuantityTakeoff,
    cost: CostBreakdown,
    project_info: dict,
    output_path: Path,
) -> Path:
    """
    Export complete design data to JSON for pyRevit consumption.
    """
    data = {
        'project': project_info,
        'quantity': takeoff.to_dict(),
        'cost': cost.to_dict(),
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return output_path


# ═══════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════

# Styles
HEADER_FONT = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF') if HAS_OPENPYXL else None
HEADER_FILL = PatternFill('solid', fgColor='4472C4') if HAS_OPENPYXL else None
TITLE_FONT = Font(name='Microsoft JhengHei', size=14, bold=True) if HAS_OPENPYXL else None
NUMBER_FORMAT = '#,##0.0'
CURRENCY_FORMAT = '#,##0_);(#,##0)'


def _style_header(ws, row: int, n_cols: int):
    """Apply header style to a row."""
    if not HAS_OPENPYXL:
        return
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def export_excel(
    takeoff: QuantityTakeoff,
    cost: CostBreakdown,
    project_info: dict,
    family_inventory: dict,
    output_path: Path,
    scwb_summary=None,
) -> Path:
    """
    Export complete report to multi-sheet Excel.

    Args:
        scwb_summary: optional SCWBSummary — adds Sheet 6 (強柱弱梁檢核)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _build_overview_sheet(wb, project_info)
    _build_concrete_sheet(wb, takeoff)
    _build_rebar_sheet(wb, takeoff)
    _build_family_sheet(wb, family_inventory)
    _build_cost_sheet(wb, cost, project_info.get('total_floor_area_m2', 0))
    if scwb_summary is not None:
        _build_scwb_sheet(wb, scwb_summary)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _build_overview_sheet(wb, info: dict):
    """Sheet 1: Design Overview."""
    ws = wb.create_sheet('1_設計總覽')

    ws['A1'] = '結構設計總覽'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    rows = [
        ('專案名稱', info.get('project_name', '')),
        ('工址位置', info.get('location', '')),
        ('總樓層', info.get('total_floors', '')),
        ('總高度 (m)', info.get('total_height_m', '')),
        ('總樓地板面積 (m²)', info.get('total_floor_area_m2', '')),
        ('結構系統', info.get('structure_system', 'RC')),
        ('設計風速 (m/s)', info.get('design_wind_speed', '')),
        ('SDS', info.get('SDS', '')),
        ('SD1', info.get('SD1', '')),
        ('用途係數 I', info.get('importance_factor', 1.0)),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def _build_concrete_sheet(wb, takeoff: QuantityTakeoff):
    """Sheet 2: Concrete Volumes."""
    ws = wb.create_sheet('2_混凝土用量')

    ws['A1'] = '混凝土用量分析 (m³)'
    ws['A1'].font = TITLE_FONT

    headers = ['構件類型', "fc'=210", "fc'=245", "fc'=280", "fc'=350", "fc'=420", '小計']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, len(headers))

    member_labels = {
        'columns':         '柱',
        'beams':           '梁',
        'slabs':           '樓板',
        'shear_walls':     '剪力牆',
        'diaphragm_walls': '連續壁',
    }

    row = 4
    for key, label in member_labels.items():
        if key not in takeoff.concrete_by_member:
            continue
        cbf = takeoff.concrete_by_member[key]
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=cbf.fc_210).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=3, value=cbf.fc_245).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=4, value=cbf.fc_280).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=5, value=cbf.fc_350).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=6, value=cbf.fc_420).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=7, value=cbf.total).number_format = NUMBER_FORMAT
        row += 1

    # Totals
    ws.cell(row=row, column=1, value='總計 m³').font = Font(bold=True)
    by_fc = takeoff.concrete_by_fc
    ws.cell(row=row, column=2, value=by_fc.fc_210).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=3, value=by_fc.fc_245).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=4, value=by_fc.fc_280).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=5, value=by_fc.fc_350).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=6, value=by_fc.fc_420).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=7, value=by_fc.total).number_format = NUMBER_FORMAT
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='FFE699')

    for i, w in enumerate([15, 12, 12, 12, 12, 12, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def _build_rebar_sheet(wb, takeoff: QuantityTakeoff):
    """Sheet 3: Rebar Tonnage."""
    ws = wb.create_sheet('3_鋼筋用量')

    ws['A1'] = '鋼筋用量分析 (噸)'
    ws['A1'].font = TITLE_FONT

    headers = ['構件類型', '配筋率假設', '混凝土量 m³', '鋼筋估算 噸']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, len(headers))

    rows = [
        ('柱',     '2.5%', takeoff.concrete_by_member['columns'].total, takeoff.rebar.columns),
        ('大梁',   '1.8%', takeoff.concrete_by_member['beams'].total, takeoff.rebar.main_beams),
        ('樓板',   '0.6%', takeoff.concrete_by_member['slabs'].total, takeoff.rebar.slabs),
        ('剪力牆', '0.8%', takeoff.concrete_by_member['shear_walls'].total, takeoff.rebar.shear_walls),
        ('連續壁', '1.2%', takeoff.concrete_by_member['diaphragm_walls'].total, takeoff.rebar.diaphragm_walls),
    ]

    for i, (label, ratio, vol, tons) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=ratio)
        ws.cell(row=i, column=3, value=vol).number_format = NUMBER_FORMAT
        ws.cell(row=i, column=4, value=tons).number_format = NUMBER_FORMAT

    end_row = 4 + len(rows)
    ws.cell(row=end_row, column=1, value='總計').font = Font(bold=True)
    ws.cell(row=end_row, column=4, value=takeoff.rebar.total).number_format = NUMBER_FORMAT
    for col in range(1, 5):
        ws.cell(row=end_row, column=col).fill = PatternFill('solid', fgColor='FFE699')

    # Density check
    ws.cell(row=end_row + 2, column=1, value='平均鋼筋密度 (kg/m³ 混凝土)').font = Font(bold=True)
    ws.cell(row=end_row + 2, column=4, value=takeoff.rebar_density_kg_m3).number_format = NUMBER_FORMAT
    ws.cell(row=end_row + 3, column=1, value='典型RC高層: 120~180 kg/m³').font = Font(italic=True, color='666666')

    for i, w in enumerate([15, 14, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def _build_family_sheet(wb, families: dict):
    """Sheet 4: Revit Family/Type Inventory."""
    ws = wb.create_sheet('4_構件清單')

    ws['A1'] = 'Revit Family/Type 清單'
    ws['A1'].font = TITLE_FONT

    ws['A3'] = '提供 pyRevit 腳本建立 Family Type 用'
    ws['A3'].font = Font(italic=True, color='666666')

    headers = ['類別', 'Family/Type', '尺寸', '數量', '備註']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=h)
    _style_header(ws, 5, len(headers))

    row = 6
    for category, items in families.items():
        for item in items:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=item.get('type', ''))
            ws.cell(row=row, column=3, value=item.get('size', ''))
            ws.cell(row=row, column=4, value=item.get('count', 0))
            ws.cell(row=row, column=5, value=item.get('note', ''))
            row += 1

    for i, w in enumerate([12, 30, 25, 10, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def _build_cost_sheet(wb, cost: CostBreakdown, total_floor_area_m2: float):
    """Sheet 5: Cost Breakdown."""
    ws = wb.create_sheet('5_造價估算')

    ws['A1'] = '結構工程造價估算 (NTD)'
    ws['A1'].font = TITLE_FONT

    ws['A3'] = '※ 本造價為設計初期參考值，實際以工程預算書為準'
    ws['A3'].font = Font(italic=True, color='AA6644')

    headers = ['項目', '說明', '金額 (NTD)', '佔比']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=h)
    _style_header(ws, 5, len(headers))

    items = [
        ('混凝土', '依fc\' 分區', cost.concrete_total),
        ('鋼筋', 'SD420W 含工料', cost.rebar),
        ('鋼骨', '含防火被覆', cost.steel),
        ('連續壁', '含設備施工', cost.diaphragm_wall),
        ('模板', '一般模板', cost.formwork),
    ]

    total = cost.grand_total
    row = 6
    for label, note, amount in items:
        if amount == 0:
            continue
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=note)
        ws.cell(row=row, column=3, value=amount).number_format = CURRENCY_FORMAT
        pct = amount / total if total > 0 else 0
        ws.cell(row=row, column=4, value=pct).number_format = '0.0%'
        row += 1

    ws.cell(row=row, column=1, value='合計').font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=total).number_format = CURRENCY_FORMAT
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='FFD966')

    if total_floor_area_m2 > 0:
        per_m2 = total / total_floor_area_m2
        ws.cell(row=row + 2, column=1, value='單位造價 (NTD/m²)').font = Font(bold=True)
        ws.cell(row=row + 2, column=3, value=per_m2).number_format = CURRENCY_FORMAT
        ws.cell(row=row + 3, column=1, value='單位造價 (NTD/坪)').font = Font(bold=True)
        ws.cell(row=row + 3, column=3, value=per_m2 * 3.30578).number_format = CURRENCY_FORMAT

    for i, w in enumerate([15, 25, 18, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def _build_scwb_sheet(wb, scwb):
    """Sheet 6: 強柱弱梁 (SCWB) check."""
    ws = wb.create_sheet('6_強柱弱梁檢核')

    ws['A1'] = '強柱弱梁檢核 (ΣMnc ≥ 6/5 ΣMnb)'
    ws['A1'].font = TITLE_FONT

    ws['A3'] = '※ 設計初期粗估：純彎曲 Mn，假設配筋率（柱2.5% / 梁1.8%），未計軸力'
    ws['A3'].font = Font(italic=True, color='AA6644')
    ws['A4'] = '※ 不替代結構技師正式 interaction-diagram 分析與簽證'
    ws['A4'].font = Font(italic=True, color='AA6644')

    # Summary block
    summary = [
        ('總接頭數', scwb.total),
        ('通過', scwb.passing),
        ('未通過', scwb.failing),
        ('通過率', f'{scwb.pass_rate * 100:.1f}%'),
        ('未通過樓層', ', '.join(str(f) for f in scwb.failing_floors) or '無'),
    ]
    for i, (label, value) in enumerate(summary, start=6):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Per-joint table
    header_row = 12
    headers = ['格點', '樓層', 'ΣMnc (kN·m)', 'ΣMnb (kN·m)', '比值', '判定']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header(ws, header_row, len(headers))

    row = header_row + 1
    for j in scwb.joints:
        ws.cell(row=row, column=1, value=j.grid)
        ws.cell(row=row, column=2, value=j.floor)
        ws.cell(row=row, column=3, value=round(j.sum_Mnc_kNm, 1)).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=4, value=round(j.sum_Mnb_kNm, 1)).number_format = NUMBER_FORMAT
        ratio_cell = ws.cell(row=row, column=5,
                             value=(round(j.ratio, 3) if j.ratio != float('inf') else '∞'))
        verdict = ws.cell(row=row, column=6, value='✓ 通過' if j.passes else '✗ 未通過')
        if not j.passes:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='FFC7CE')
        else:
            verdict.font = Font(color='008000')
        row += 1

    for i, w in enumerate([10, 8, 16, 16, 10, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


# ═══════════════════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════════════════

# Traditional-Chinese CID font shipped with reportlab (Adobe-CNS1)
_CJK_FONT = 'MSung-Light'
_cjk_registered = False


def _ensure_cjk_font():
    """Register the built-in Traditional-Chinese CID font once."""
    global _cjk_registered
    if not _cjk_registered:
        pdfmetrics.registerFont(UnicodeCIDFont(_CJK_FONT))
        _cjk_registered = True


def _pdf_styles():
    """Build paragraph styles using the CJK font."""
    styles = getSampleStyleSheet()
    title = ParagraphStyle('CJKTitle', parent=styles['Title'],
                           fontName=_CJK_FONT, fontSize=18, spaceAfter=6)
    heading = ParagraphStyle('CJKHeading', parent=styles['Heading2'],
                              fontName=_CJK_FONT, fontSize=13, spaceBefore=12, spaceAfter=4)
    body = ParagraphStyle('CJKBody', parent=styles['Normal'],
                           fontName=_CJK_FONT, fontSize=9, leading=13)
    note = ParagraphStyle('CJKNote', parent=body, fontSize=8,
                           textColor=colors.HexColor('#AA6644'))
    return {'title': title, 'heading': heading, 'body': body, 'note': note}


def _pdf_table(data, col_widths=None, highlight_rows=None):
    """Build a styled reportlab Table. Row 0 is the header."""
    t = Table(data, colWidths=col_widths)
    style = [
        ('FONTNAME', (0, 0), (-1, -1), _CJK_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    for r in (highlight_rows or []):
        style.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FFC7CE')))
    t.setStyle(TableStyle(style))
    return t


def export_pdf(
    takeoff: QuantityTakeoff,
    cost: CostBreakdown,
    project_info: dict,
    output_path: Path,
    scwb_summary=None,
) -> Path:
    """Export a summary PDF report (A4, Traditional Chinese).

    Args:
        scwb_summary: optional SCWBSummary — adds a 強柱弱梁 section
    """
    if not HAS_REPORTLAB:
        raise ImportError(
            "reportlab is required for PDF export. Install: pip install reportlab")

    _ensure_cjk_font()
    st = _pdf_styles()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=18 * mm, rightMargin=18 * mm,
        title='結構設計初步報告',
    )
    flow = []

    # ── Title ──
    flow.append(Paragraph('結構設計初步報告', st['title']))
    flow.append(Paragraph(
        '設計初期用量、造價與耐震粗估 — 不替代結構技師正式計算與簽證', st['note']))
    flow.append(Spacer(1, 8))

    # ── 1. Project overview ──
    flow.append(Paragraph('1　專案資訊', st['heading']))
    ov = [
        ['項目', '內容'],
        ['專案名稱', str(project_info.get('project_name', ''))],
        ['工址位置', str(project_info.get('location', ''))],
        ['結構系統', str(project_info.get('structure_system', 'RC'))],
        ['地上樓層', str(project_info.get('total_floors', ''))],
        ['總高度 (m)', str(project_info.get('total_height_m', ''))],
        ['總樓地板 (m²)', f"{project_info.get('total_floor_area_m2', 0):,.0f}"],
        ['設計風速 (m/s)', str(project_info.get('design_wind_speed', ''))],
        ['SDS / SD1', f"{project_info.get('SDS', '')} / {project_info.get('SD1', '')}"],
    ]
    flow.append(_pdf_table(ov, col_widths=[55 * mm, 110 * mm]))

    # ── 2. Concrete ──
    flow.append(Paragraph('2　混凝土用量 (m³)', st['heading']))
    by_fc = takeoff.concrete_by_fc
    conc = [['fc\' 強度', '體積 (m³)']]
    for fc_key in ('fc_210', 'fc_245', 'fc_280', 'fc_350', 'fc_420'):
        conc.append([fc_key.replace('fc_', "fc'="), f"{getattr(by_fc, fc_key):,.1f}"])
    conc.append(['總計', f"{by_fc.total:,.1f}"])
    flow.append(_pdf_table(conc, col_widths=[55 * mm, 55 * mm]))

    # ── 3. Rebar ──
    flow.append(Paragraph('3　鋼筋用量 (噸)', st['heading']))
    rb = takeoff.rebar
    rebar = [
        ['構件', '鋼筋 (噸)'],
        ['柱', f'{rb.columns:,.1f}'],
        ['大梁', f'{rb.main_beams:,.1f}'],
        ['樓板', f'{rb.slabs:,.1f}'],
        ['剪力牆', f'{rb.shear_walls:,.1f}'],
        ['連續壁', f'{rb.diaphragm_walls:,.1f}'],
        ['總計', f'{rb.total:,.1f}'],
    ]
    flow.append(_pdf_table(rebar, col_widths=[55 * mm, 55 * mm]))
    density = takeoff.rebar_density_kg_m3
    note = f'平均鋼筋密度：{density:,.1f} kg/m³（典型 RC 高層 120–180 kg/m³）'
    if density < 120:
        note += ' ⚠ 偏低'
    elif density > 180:
        note += ' ⚠ 偏高'
    flow.append(Spacer(1, 3))
    flow.append(Paragraph(note, st['note']))

    # ── 4. Cost ──
    flow.append(Paragraph('4　造價估算 (NTD)', st['heading']))
    cd = cost.to_dict()
    cost_rows = [['項目', '金額 (NTD)']]
    for label, key in [('混凝土', 'concrete_total'), ('鋼筋', 'rebar'),
                       ('鋼骨', 'steel'), ('連續壁', 'diaphragm_wall'),
                       ('模板', 'formwork')]:
        cost_rows.append([label, f"{cd[key]:,.0f}"])
    cost_rows.append(['合計', f"{cd['grand_total']:,.0f}"])
    flow.append(_pdf_table(cost_rows, col_widths=[55 * mm, 55 * mm]))
    area = project_info.get('total_floor_area_m2', 0)
    if area > 0:
        per_m2 = cost.grand_total / area
        flow.append(Spacer(1, 3))
        flow.append(Paragraph(
            f'單位造價：NT$ {per_m2:,.0f}/m²（NT$ {per_m2 * 3.30578:,.0f}/坪）',
            st['note']))

    # ── 5. SCWB ──
    if scwb_summary is not None:
        flow.append(Paragraph('5　強柱弱梁檢核 (ΣMnc ≥ 6/5 ΣMnb)', st['heading']))
        s = scwb_summary
        flow.append(Paragraph(
            f'總接頭 {s.total}　通過 {s.passing}　未通過 {s.failing}　'
            f'通過率 {s.pass_rate * 100:.1f}%', st['body']))
        if s.failing_floors:
            flow.append(Paragraph(
                '未通過樓層：' + ', '.join(str(f) for f in s.failing_floors),
                st['note']))
        # Show only failing joints (or first 20 if all pass)
        failing = [j for j in s.joints if not j.passes]
        show = failing if failing else s.joints[:20]
        if show:
            jt = [['格點', '樓層', 'ΣMnc', 'ΣMnb', '比值', '判定']]
            for j in show:
                ratio = '∞' if j.ratio == float('inf') else f'{j.ratio:.2f}'
                jt.append([j.grid, str(j.floor), f'{j.sum_Mnc_kNm:,.0f}',
                           f'{j.sum_Mnb_kNm:,.0f}', ratio,
                           '✓' if j.passes else '✗'])
            hl = [i + 1 for i, j in enumerate(show) if not j.passes]
            flow.append(Spacer(1, 3))
            flow.append(_pdf_table(jt, highlight_rows=hl))
        flow.append(Spacer(1, 3))
        flow.append(Paragraph(
            '※ 設計初期粗估：純彎曲 Mn、假設配筋率、未計軸力，'
            '不替代結構技師簽證', st['note']))

    doc.build(flow)
    return output_path
